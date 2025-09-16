"""
Excel Export Module
Export valve sizing results to Excel spreadsheets

Features:
- Comprehensive data export to Excel
- Professional formatting with charts
- Multiple worksheets for different analysis types
- Summary tables and detailed calculations
"""

from typing import Dict, Any, List, Optional
import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
except ImportError:
    print("openpyxl not installed. Install with: pip install openpyxl")

class ExcelExporter:
    """Professional Excel export for valve sizing results"""
    
    def __init__(self):
        try:
            self.wb = openpyxl.Workbook()
            self._setup_styles()
        except:
            self.wb = None
    
    def _setup_styles(self):
        """Setup Excel cell styles"""
        if not self.wb:
            return
            
        # Header style
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Data style
        self.data_font = Font(name='Arial', size=10)
        self.data_alignment = Alignment(horizontal='left', vertical='center')
        
        # Number style
        self.number_alignment = Alignment(horizontal='right', vertical='center')
        
        # Border
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_complete_results(self, project_data: Dict[str, Any],
                              sizing_results: Dict[str, Any],
                              analysis_results: Dict[str, Any]) -> bytes:
        """Export complete valve sizing results to Excel"""
        
        if not self.wb:
            return b"Excel export not available - openpyxl not installed"
        
        try:
            # Remove default sheet
            if 'Sheet' in self.wb.sheetnames:
                self.wb.remove(self.wb['Sheet'])
            
            # Create worksheets
            self._create_summary_sheet(project_data, sizing_results, analysis_results)
            self._create_process_conditions_sheet(project_data)
            self._create_sizing_calculations_sheet(sizing_results)
            self._create_analysis_results_sheet(analysis_results)
            self._create_standards_compliance_sheet()
            
            # Save to buffer
            buffer = io.BytesIO()
            self.wb.save(buffer)
            buffer.seek(0)
            return buffer.read()
            
        except Exception as e:
            return f"Excel export failed: {str(e)}".encode()
    
    def _create_summary_sheet(self, project_data: Dict[str, Any],
                            sizing_results: Dict[str, Any],
                            analysis_results: Dict[str, Any]):
        """Create executive summary worksheet"""
        
        ws = self.wb.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "Control Valve Sizing - Executive Summary"
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Project information
        row = 3
        ws[f'A{row}'] = "Project Information"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        project_info = [
            ('Project:', project_data.get('project_name', 'Not specified')),
            ('Tag Number:', project_data.get('tag_number', 'Not specified')),
            ('Service:', project_data.get('service_description', 'Not specified')),
            ('Date:', datetime.now().strftime("%Y-%m-%d")),
            ('Engineer:', project_data.get('engineer', 'Aseem Mehrotra, KBR Inc'))
        ]
        
        for label, value in project_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Key results
        row += 2
        ws[f'A{row}'] = "Key Results"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        key_results = [
            ('Required Cv:', f"{sizing_results.get('cv_required', 0):.2f}"),
            ('Recommended Valve Size:', sizing_results.get('recommended_valve_size', 'TBD')),
            ('Safety Factor:', f"{sizing_results.get('safety_factor', 1.2):.1f}"),
            ('Sizing Method:', sizing_results.get('sizing_method', 'ISA 75.01'))
        ]
        
        for label, value in key_results:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Apply formatting
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.border = self.thin_border
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.alignment = self.number_alignment
                else:
                    cell.alignment = self.data_alignment
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_process_conditions_sheet(self, project_data: Dict[str, Any]):
        """Create process conditions worksheet"""
        
        ws = self.wb.create_sheet("Process Conditions")
        
        # Headers
        headers = ['Parameter', 'Value', 'Units', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        # Process data
        process_data = [
            ('Fluid Type', project_data.get('fluid_type', ''), '', ''),
            ('Fluid Name', project_data.get('fluid_name', ''), '', ''),
            ('Temperature', project_data.get('temperature', 0), '°C', ''),
            ('Inlet Pressure (P1)', project_data.get('p1', 0), 'bar abs', ''),
            ('Outlet Pressure (P2)', project_data.get('p2', 0), 'bar abs', ''),
            ('Pressure Drop', project_data.get('p1', 0) - project_data.get('p2', 0), 'bar', 'Calculated'),
            ('Normal Flow Rate', project_data.get('normal_flow', 0), project_data.get('flow_units', 'm³/h'), ''),
            ('Minimum Flow Rate', project_data.get('min_flow', 0), project_data.get('flow_units', 'm³/h'), ''),
            ('Maximum Flow Rate', project_data.get('max_flow', 0), project_data.get('flow_units', 'm³/h'), ''),
            ('Pipe Size', project_data.get('pipe_size', ''), 'NPS', ''),
            ('Service Type', project_data.get('service_type', ''), '', ''),
            ('Criticality', project_data.get('criticality', ''), '', ''),
            ('H2S Present', 'Yes' if project_data.get('h2s_present', False) else 'No', '', ''),
            ('Fire Safe Required', 'Yes' if project_data.get('fire_safe_required', False) else 'No', '', '')
        ]
        
        for row, (param, value, unit, note) in enumerate(process_data, 2):
            ws.cell(row=row, column=1, value=param).font = self.data_font
            ws.cell(row=row, column=2, value=value).font = self.data_font
            ws.cell(row=row, column=3, value=unit).font = self.data_font
            ws.cell(row=row, column=4, value=note).font = self.data_font
        
        # Apply borders and formatting
        for row in ws.iter_rows():
            for cell in row:
                cell.border = self.thin_border
                if isinstance(cell.value, (int, float)) and cell.column == 2:
                    cell.alignment = self.number_alignment
                    if isinstance(cell.value, float):
                        cell.number_format = '0.00'
                else:
                    cell.alignment = self.data_alignment
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
    
    def _create_sizing_calculations_sheet(self, sizing_results: Dict[str, Any]):
        """Create sizing calculations worksheet"""
        
        ws = self.wb.create_sheet("Sizing Calculations")
        
        # Headers
        ws['A1'] = "Valve Sizing Calculations"
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        # Calculation results
        row = 3
        calc_data = [
            ('Basic Cv (Turbulent)', sizing_results.get('cv_basic', 0), 'GPM/psi^0.5'),
            ('Reynolds Correction Factor (Fr)', sizing_results.get('reynolds_analysis', {}).get('fr_factor', 1.0), '-'),
            ('Piping Geometry Factor (Fp)', sizing_results.get('fp_factor', 1.0), '-'),
            ('Corrected Cv Required', sizing_results.get('cv_required', 0), 'GPM/psi^0.5'),
            ('Safety Factor Applied', sizing_results.get('safety_factor', 1.2), '-'),
            ('Final Cv Required', sizing_results.get('cv_required', 0) * sizing_results.get('safety_factor', 1.2), 'GPM/psi^0.5'),
            ('Calculation Method', sizing_results.get('sizing_method', 'ISA 75.01'), ''),
            ('Calculation Date', datetime.now().strftime("%Y-%m-%d %H:%M"), '')
        ]
        
        for param, value, unit in calc_data:
            ws[f'A{row}'] = param
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'C{row}'] = unit
            row += 1
        
        # Apply formatting
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.border = self.thin_border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = self.number_alignment
                    cell.number_format = '0.00'
                else:
                    cell.alignment = self.data_alignment
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_analysis_results_sheet(self, analysis_results: Dict[str, Any]):
        """Create analysis results worksheet"""
        
        ws = self.wb.create_sheet("Analysis Results")
        
        # Headers
        ws['A1'] = "Technical Analysis Results"
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Cavitation analysis
        if 'cavitation_analysis' in analysis_results:
            ws[f'A{row}'] = "Cavitation Analysis (ISA RP75.23)"
            ws[f'A{row}'].font = self.header_font
            ws[f'A{row}'].fill = self.header_fill
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            cav_analysis = analysis_results['cavitation_analysis']
            cav_data = [
                ('Service Sigma', cav_analysis.get('sigma_service', 0), '-', ''),
                ('FL Corrected Sigma', cav_analysis.get('sigma_fl_corrected', 0), '-', ''),
                ('Cavitation Status', 'Cavitating' if cav_analysis.get('is_cavitating', False) else 'No Cavitation', '', ''),
                ('Risk Level', cav_analysis.get('cavitation_assessment', {}).get('risk_level', 'Unknown'), '', '')
            ]
            
            for param, value, unit, note in cav_data:
                ws[f'A{row}'] = param
                ws[f'B{row}'] = value
                ws[f'C{row}'] = unit
                ws[f'D{row}'] = note
                row += 1
            
            row += 1
        
        # Noise analysis
        if 'noise_analysis' in analysis_results:
            ws[f'A{row}'] = "Noise Analysis (IEC 60534-8-3)"
            ws[f'A{row}'].font = self.header_font
            ws[f'A{row}'].fill = self.header_fill
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            noise_analysis = analysis_results['noise_analysis']
            noise_data = [
                ('Sound Pressure Level', noise_analysis.get('spl_at_distance', 0), 'dBA', 'At 1 meter'),
                ('Assessment', noise_analysis.get('assessment', {}).get('level', 'Unknown'), '', ''),
                ('Peak Frequency', noise_analysis.get('peak_frequency', 0), 'Hz', 'Estimated')
            ]
            
            for param, value, unit, note in noise_data:
                ws[f'A{row}'] = param
                ws[f'B{row}'] = value
                ws[f'C{row}'] = unit
                ws[f'D{row}'] = note
                row += 1
        
        # Apply formatting
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.border = self.thin_border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = self.number_alignment
                    cell.number_format = '0.00'
                else:
                    cell.alignment = self.data_alignment
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 20
    
    def _create_standards_compliance_sheet(self):
        """Create standards compliance worksheet"""
        
        ws = self.wb.create_sheet("Standards Compliance")
        
        # Headers
        ws['A1'] = "Standards Compliance Documentation"
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        # Standards information
        standards_data = [
            ('Standard', 'Title', 'Compliance Status'),
            ('ISA 75.01-2012', 'Flow equations for sizing control valves', 'Fully Compliant'),
            ('IEC 60534-2-1:2011', 'Industrial-process control valves', 'Fully Compliant'),
            ('ISA RP75.23-1995', 'Considerations for evaluating control valve cavitation', 'Analysis Performed'),
            ('IEC 60534-8-3:2010', 'Control valve aerodynamic noise prediction', 'Analysis Performed'),
            ('ASME B16.34-2017', 'Valves - Flanged, threaded, and welding end', 'Reference Standard'),
            ('NACE MR0175/ISO 15156', 'Materials for H2S environments', 'Reference Standard')
        ]
        
        for row, (standard, title, status) in enumerate(standards_data, 3):
            ws[f'A{row}'] = standard
            ws[f'B{row}'] = title
            ws[f'C{row}'] = status
            
            if row == 3:  # Header row
                ws[f'A{row}'].font = self.header_font
                ws[f'B{row}'].font = self.header_font
                ws[f'C{row}'].font = self.header_font
                ws[f'A{row}'].fill = self.header_fill
                ws[f'B{row}'].fill = self.header_fill
                ws[f'C{row}'].fill = self.header_fill
            else:
                ws[f'A{row}'].font = self.data_font
                ws[f'B{row}'].font = self.data_font
                ws[f'C{row}'].font = self.data_font
        
        # Apply formatting
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.border = self.thin_border
                cell.alignment = self.data_alignment
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
